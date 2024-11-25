import csv
import json
import mimetypes
import os
import shutil
import subprocess
import zipfile
from contextlib import asynccontextmanager

import aiofiles
import geojson
import openpyxl
import psycopg2
import requests
from fastapi import FastAPI, File, HTTPException, UploadFile, status
from shapely import wkb, wkt

from api.version import __version__

DB_HOST = "localhost"
DB_PORT = 5432
DB_USER = "postgres"
DB_PASSWORD = "postgres"
DB_NAME = "data"
DEFAULT_CHUNK_SIZE = 1024 * 1024 * 50  # 50 megabytes
GEOGRAPHIES = []


@asynccontextmanager
async def lifespan(app: FastAPI):
    """
    FastAPI lifespan event to initialize the application.

    This event is called on application startup and shutdown. On startup, it reads the
    geographies.json file to initialize the list of geographies.
    """
    global GEOGRAPHIES
    with open("api/geographies.json", "r") as f:
        GEOGRAPHIES = json.loads(f.read())
    yield


app = FastAPI(
    title="pg-upload-anything-api",
    description="API for uploading any geographic data to a PostgreSQL database.",
    lifespan=lifespan,
    version=__version__,
)


def find_matching_geographies(column_names: list):
    """
    Finds and returns a list of geographies that match the provided column names.

    Iterates through the predefined `GEOGRAPHIES` list, checking if the given
    column names match the potential names for each field in a geography. A geography
    is considered a match if all its fields have at least one column name that matches
    one of their potential names.

    Args:
        column_names (list): A list of column names to match against the potential names of geography fields.

    Returns:
        list: A list of geographies that have fields matching the provided column names. Each geography
              includes a `field_matches` dictionary indicating which column names matched each field.
    """
    matching_geographies = []

    for geography in GEOGRAPHIES:
        valid_match = []
        field_matches = {}
        for column in geography["fields"]:
            valid_field_match = False
            for geo_column in geography["fields"][column]["potential_names"]:
                for column_name in column_names:
                    if column_name.strip() == geo_column:
                        valid_field_match = True
                        field_matches[column] = column_name.strip()
            valid_match.append(valid_field_match)
        if all(valid_match):
            geography["field_matches"] = field_matches
            matching_geographies.append(geography)

    return matching_geographies


def import_point_dataset(
    file_path: str,
    latitude: str,
    longitude: str,
    table_name: str,
):
    """
    Imports a point dataset into a PostgreSQL database using the ogr2ogr command.

    This function utilizes ogr2ogr to convert a file containing point data into a PostgreSQL
    table. The function specifies the latitude and longitude fields to be used for point
    geometry creation.

    Args:
        file_path (str): The path to the file containing the point dataset.
        latitude (str): The column name in the file that represents the latitude values.
        longitude (str): The column name in the file that represents the longitude values.
        table_name (str): The name of the PostgreSQL table where the data will be imported.

    """
    subprocess.call(
        f"""ogr2ogr \
        -f "PostgreSQL" PG:"dbname={DB_NAME} user={DB_USER} password={DB_PASSWORD} host={DB_HOST}" \
        {file_path} \
        -oo X_POSSIBLE_NAMES={longitude}* \
        -oo Y_POSSIBLE_NAMES={latitude}* \
        -nln {table_name} \
        -a_srs "EPSG:4326" \
        -lco GEOMETRY_NAME=geom \
        -lco FID=gid \
        -nlt POINT \
        -overwrite""",
        shell=True,
    )


def join_to_map_service(
    file_path: str,
    table_name: str,
    map_name: str,
    table_match_column: str,
    map_match_column: str,
):
    """
    Joins a file containing data to a map service using the ogr2ogr command.

    This function utilizes ogr2ogr to convert a file containing data into a PostgreSQL
    table and then joins that table to a map service table based on matching values between
    the two tables.

    Args:
        file_path (str): The path to the file containing the data.
        table_name (str): The name of the PostgreSQL table where the data will be imported.
        map_name (str): The name of the map service table that the data will be joined to.
        table_match_column (str): The column name in the file that will be used to join the data to the map service.
        map_match_column (str): The column name in the map service table that will be used to join the data to the map service.
    """
    subprocess.call(
        f"""ogr2ogr -f "PostgreSQL" PG:"dbname={DB_NAME} user={DB_USER} password={DB_PASSWORD} host={DB_HOST}" \
        {file_path} -nln {table_name}_temp -lco FID=gid  -overwrite""",
        shell=True,
    )
    drop_table_query = f"""DROP TABLE IF EXISTS "{table_name}";"""
    join_sql = f"""CREATE TABLE "{table_name}" AS
            SELECT a.*, b."{map_match_column}", b.geom
            FROM "{table_name}_temp" as a
            LEFT JOIN "{map_name}" as b
            ON LOWER(a."{table_match_column}") = LOWER(b."{map_match_column}");
        """
    drop_temp_table_query = f"""DROP TABLE IF EXISTS "{table_name}_temp";"""

    connection = psycopg2.connect(
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        host=DB_HOST,
    )
    cursor = connection.cursor()
    cursor.execute(drop_table_query)
    cursor.execute(join_sql)
    cursor.execute(drop_temp_table_query)
    connection.commit()
    cursor.close()
    connection.close()


def upload_geographic_file(file_path: str, table_name: str):
    """
    Uploads a geographic file to a PostgreSQL database using the ogr2ogr command.

    This function utilizes ogr2ogr to import a geographic file into a PostgreSQL
    table. The function assigns a geometry name and FID to the table and overwrites
    any existing table with the same name.

    Args:
        file_path (str): The path to the geographic file to be uploaded.
        table_name (str): The name of the PostgreSQL table where the data will be stored.
    """
    subprocess.call(
        f"""ogr2ogr -f "PostgreSQL" PG:"dbname={DB_NAME} user={DB_USER} password={DB_PASSWORD} host={DB_HOST}" \
        {file_path} -nln {table_name} -lco FID=gid -lco GEOMETRY_NAME=geom  -overwrite""",
        shell=True,
    )


def convert_csv_to_geojson(file_path: str):
    """
    Converts a CSV file to a GeoJSON file.

    Args:
        file_path (str): The path to the CSV file to be converted.
    """
    geojson_object = {"type": "FeatureCollection", "features": []}
    with open(file_path, "r") as f:
        reader = csv.DictReader(f)
        for row in reader:
            geojson_object["features"].append(
                {
                    "type": "Feature",
                    "geometry": json.loads(row["geojson"]),
                    "properties": {k: v for k, v in row.items() if k != "geojson"},
                }
            )
    with open(file_path.replace(".csv", ".geojson"), "w") as f:
        json.dump(geojson_object, f)


def convert_wkt_geometry_to_geojson(wkt_string: str):
    """
    Converts a WKT geometry to a GeoJSON geometry.

    Args:
        wkt_string (str): The WKT geometry to be converted.

    Returns:
        dict: A GeoJSON geometry.
    """

    geometry = wkt.loads(wkt_string)

    geojson_string = geojson.dumps(geometry.__geo_interface__)

    return json.loads(geojson_string)


def convert_wkt_to_geojson(file_path: str):
    """
    Converts a CSV file with WKT geometries to a GeoJSON file.

    Args:
        file_path (str): The path to the CSV file to be converted.
    """
    geojson_object = {"type": "FeatureCollection", "features": []}
    with open(file_path, "r") as f:
        reader = csv.DictReader(f)
        for row in reader:
            geojson_object["features"].append(
                {
                    "type": "Feature",
                    "geometry": convert_wkt_geometry_to_geojson(row["wkt"]),
                    "properties": {k: v for k, v in row.items() if k != "wkt"},
                }
            )
    with open(file_path.replace(".csv", ".geojson"), "w") as f:
        json.dump(geojson_object, f)


def convert_wkb_geometry_to_geojson(wkb_string: str):
    """
    Converts a WKB geometry to a GeoJSON geometry.

    Args:
        wkb_string (str): The WKB geometry to be converted.

    Returns:
        dict: A GeoJSON geometry.
    """

    geometry = wkb.loads(wkb_string, hex=True)

    geojson_string = geojson.dumps(geometry.__geo_interface__)

    return json.loads(geojson_string)


def convert_wkb_to_geojson(file_path: str):
    """
    Converts a CSV file with WKB geometries to a GeoJSON file.

    Args:
        file_path (str): The path to the CSV file to be converted.
    """
    geojson_object = {"type": "FeatureCollection", "features": []}
    with open(file_path, "r") as f:
        reader = csv.DictReader(f)
        for row in reader:
            geojson_object["features"].append(
                {
                    "type": "Feature",
                    "geometry": convert_wkb_geometry_to_geojson(row["wkb"]),
                    "properties": {k: v for k, v in row.items() if k != "wkb"},
                }
            )
    with open(file_path.replace(".csv", ".geojson"), "w") as f:
        json.dump(geojson_object, f)


def upload_csv_file(write_file_path: str, file_name: str) -> object:
    """
    Uploads a CSV file to a PostgreSQL database by finding a matching geography and either importing as points or joining to a map service.

    This function reads the first two lines of the CSV file, finds a matching geography by comparing the column names against the potential names of geography fields, and either imports the file as point data or joins it to a map service table based on the matching geography.

    Args:
        write_file_path (str): The path to the CSV file to be uploaded.
        file_name (str): The name of the CSV file.

    Returns:
        object: A dictionary containing a status and message. The status is True if the upload is successful and False if no matching geography is found. The message is a human-readable description of the result.
    """
    with open(write_file_path) as input_file:
        head = [next(input_file) for _ in range(2)]
    matching_geographies = find_matching_geographies(head[0].split(","))
    if len(matching_geographies) == 0:
        return {"status": False, "message": "No matching geography found"}
    matching_geography = sorted(matching_geographies, key=lambda x: x["rank"])[0]

    if matching_geography["name"] == "latitude_and_longitude":
        import_point_dataset(
            file_path=write_file_path,
            latitude=matching_geography["field_matches"]["latitude"],
            longitude=matching_geography["field_matches"]["longitude"],
            table_name=file_name.split(".")[0],
        )

    elif matching_geography["name"] == "geojson_geometry":
        convert_csv_to_geojson(file_path=write_file_path)
        upload_geographic_file(
            file_path=write_file_path.replace(".csv", ".geojson"),
            table_name=file_name.split(".")[0],
        )
        os.remove(write_file_path.replace(".csv", ".geojson"))

    elif matching_geography["name"] == "wkt_geometry":
        convert_wkt_to_geojson(file_path=write_file_path)
        upload_geographic_file(
            file_path=write_file_path.replace(".csv", ".geojson"),
            table_name=file_name.split(".")[0],
        )
        os.remove(write_file_path.replace(".csv", ".geojson"))

    elif matching_geography["name"] == "wkb_geometry":
        convert_wkb_to_geojson(file_path=write_file_path)
        upload_geographic_file(
            file_path=write_file_path.replace(".csv", ".geojson"),
            table_name=file_name.split(".")[0],
        )
        os.remove(write_file_path.replace(".csv", ".geojson"))

    else:
        map_match_column = list(matching_geography["field_matches"].keys())[0]

        join_to_map_service(
            file_path=write_file_path,
            table_name=file_name.split(".")[0],
            map_name=matching_geography["name"],
            table_match_column=matching_geography["field_matches"][map_match_column],
            map_match_column=map_match_column,
        )

    return {
        "status": True,
    }


@app.post(path="/api/v1/upload_file")
async def upload_file(file: UploadFile = File(...)):
    """
    Upload a file to the server and import it into a PostgreSQL database.

    The file can be in any of the following formats:
    - CSV
    - Excel (xlsx)
    - GeoPackage
    - GeoJSON
    - GML
    - GPX
    - KML
    - SQLite
    - Shapefile
    - ZIP (containing one of the above formats)
    """

    valid_file_type = False

    valid_file_types = [
        "text/csv",
        "application/vnd.ms-excel",
        "application/geopackage+sqlite3",
        "application/geo+json",
        "application/gml+xml",
        "application/gpx+xml",
        "application/vnd.google-earth.kml+xml",
        "application/vnd.sqlite3",
        "application/vnd.shp",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/zip",
    ]

    if file.content_type in valid_file_types:
        valid_file_type = True

    if valid_file_type is False:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f'Please upload a valid file type. {" ,".join(valid_file_types)}',
        )

    file_name = file.filename

    write_file_path = f"{os.getcwd()}/media/{file.filename}"

    try:
        if os.path.exists(f"{os.getcwd()}/media/") is False:
            os.mkdir(f"{os.getcwd()}/media/")
        async with aiofiles.open(write_file_path, "wb") as new_file:
            while chunk := await file.read(DEFAULT_CHUNK_SIZE):
                await new_file.write(chunk)
    except Exception:
        media_directory = os.listdir(f"{os.getcwd()}/media/")
        for file in media_directory:
            if file_name in file:
                os.remove(f"{os.getcwd()}/media/{file}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="There was an error uploading the file",
        )

    if (
        file.content_type
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ):
        new_file_name = file_name.split(".")[0]
        os.makedirs(f"{os.getcwd()}/media/{new_file_name}")
        workbook = openpyxl.load_workbook(write_file_path)
        results = []
        for sheet in workbook:
            worksheet = workbook[sheet.title]
            with open(
                f"{os.getcwd()}/media/{new_file_name}/{sheet.title}.csv",
                "w",
                newline="",
            ) as csvfile:
                csvwriter = csv.writer(csvfile)
                for row in worksheet.iter_rows(values_only=True):
                    csvwriter.writerow(row)
            result = upload_csv_file(
                write_file_path=f"{os.getcwd()}/media/{new_file_name}/{sheet.title}.csv",
                file_name=sheet.title,
            )
            results.append(result["status"])
        shutil.rmtree(f"{os.getcwd()}/media/{new_file_name}")
        media_directory = os.listdir(f"{os.getcwd()}/media/")
        for file in media_directory:
            if file_name in file:
                os.remove(f"{os.getcwd()}/media/{file}")
        if True not in results:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="There was an error uploading the file",
            )
    elif file.content_type == "text/csv":
        result = upload_csv_file(write_file_path=write_file_path, file_name=file_name)

        if result["status"] is False:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=result["message"],
            )

        media_directory = os.listdir(f"{os.getcwd()}/media/")
        for file in media_directory:
            if file_name in file:
                os.remove(f"{os.getcwd()}/media/{file}")
    else:
        new_file_name = file_name.split(".")[0]
        if file.content_type == "application/zip":
            with zipfile.ZipFile(write_file_path, "r") as zip_ref:
                zip_ref.extractall(f"{os.getcwd()}/media/{new_file_name}")
        media_directory = os.listdir(f"{os.getcwd()}/media/{new_file_name}")
        valid_file_type = False
        file_extension = media_directory[0].split(".")[-1]
        valid_file_extension = ""
        for file in media_directory:
            file_path = f"{media_directory}/{file}"
            mime_type, _ = mimetypes.guess_type(file_path)
            file_extension = file_path.split(".")[-1]
            if mime_type in valid_file_types:
                valid_file_type = True
                valid_file_extension = file_path.split(".")[-1]
            if file_extension.lower() in ["gdb", "tab", "shp"]:
                valid_file_extension = file_path.split(".")[-1]
                valid_file_type = True
        if valid_file_type is False:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f'Please upload a valid file type within your zip file. {" ,".join(valid_file_types)}',
            )

        if file_extension.lower() == "csv":
            result = upload_csv_file(
                write_file_path=f"{os.getcwd()}/media/{new_file_name}/{new_file_name}.{file_extension}",
                file_name=file_name.split(".")[0],
            )

            if result["status"] is False:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=result["message"],
                )
        elif file_extension.lower() == "xlsx":
            new_file_name = file_name.split(".")[0]
            os.makedirs(f"{os.getcwd()}/media/{new_file_name}")
            workbook = openpyxl.load_workbook(write_file_path)
            results = []
            for sheet in workbook:
                worksheet = workbook[sheet.title]
                with open(
                    f"{os.getcwd()}/media/{new_file_name}/{sheet.title}.csv",
                    "w",
                    newline="",
                ) as csvfile:
                    csvwriter = csv.writer(csvfile)
                    for row in worksheet.iter_rows(values_only=True):
                        csvwriter.writerow(row)
                result = upload_csv_file(
                    write_file_path=f"{os.getcwd()}/media/{new_file_name}/{sheet.title}.csv",
                    file_name=sheet.title,
                )
                results.append(result["status"])
            shutil.rmtree(f"{os.getcwd()}/media/{new_file_name}")
            media_directory = os.listdir(f"{os.getcwd()}/media/")
            for file in media_directory:
                if file_name in file:
                    os.remove(f"{os.getcwd()}/media/{file}")
            if True not in results:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="There was an error uploading the file",
                )
        else:
            upload_geographic_file(
                file_path=f"{os.getcwd()}/media/{new_file_name}/{new_file_name}.{valid_file_extension}",
                table_name=file_name.split(".")[0],
            )
            media_directory = os.listdir(f"{os.getcwd()}/media/")
            for file in media_directory:
                if file_name in file:
                    os.remove(f"{os.getcwd()}/media/{file}")
            shutil.rmtree(f"{os.getcwd()}/media/{new_file_name}")

    return {"status": "success"}


def upload_arcgis_service(url: str, service_name: str):
    subprocess.call(
        f"""ogr2ogr -f "PostgreSQL" PG:"dbname={DB_NAME} user={DB_USER} password={DB_PASSWORD} host={DB_HOST}" \
        "{url}/query?where=1=1&outfields=*&f=geojson" -nln {service_name} -lco FID=gid -lco GEOMETRY_NAME=geom  -overwrite""",
        shell=True,
    )


def download_arcgis_service_information(url: str):
    """
    Downloads an ArcGIS service using the ArcGIS REST API and then uploads each layer
    in the service to the server.

    Args:
        url (str): The URL of the ArcGIS service to be downloaded.

    Raises:
        HTTPException: If there is an error downloading the service.
    """
    response = requests.get(f"{url}?f=pjson")

    if response.status_code != 200 or "error" in response.json():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="There was an error downloading the service",
        )

    data = response.json()

    if "layers" in data:
        if url[-1] != "/":
            url += "/"
        for layer in data["layers"]:
            layer_response = requests.get(f"{url}{layer['id']}?f=json")

            if layer_response.status_code != 200 or "error" in layer_response.json():
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="There was an error downloading the service",
                )

            layer_information = layer_response.json()

            if "Query" in layer_information["capabilities"]:
                upload_arcgis_service(
                    url=f"{url}{layer_information['id']}",
                    service_name=layer_information["name"].lower().replace(" ", "_"),
                )

    else:
        if "Query" in data["capabilities"]:
            upload_arcgis_service(
                url=url, service_name=data["name"].lower().replace(" ", "_")
            )


def upload_google_sheets(url: str):
    """
    Downloads a Google Sheets spreadsheet and imports it into a PostgreSQL database.

    The file is downloaded using the requests library and imported into a PostgreSQL database using the psycopg2 library.

    Args:
        url (str): The URL of the Google Sheets spreadsheet to be imported.
    """
    # TODO support multiple sheets
    google_doc_id = url.split("d/")[1].split("/")[0]
    url = f"https://docs.google.com/spreadsheets/d/{google_doc_id}/export?format=csv&gid=0"
    response = requests.get(url)

    if response.status_code != 200:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="There was an error downloading the spreadsheet",
        )

    file_name = (
        response.headers.get("content-disposition")
        .split("filename=")[1]
        .split(";")[0]
        .replace('"', "")
        .replace(".csv", "")
        .lower()
        .replace(" ", "_")
        .replace("-", "_")
    )

    with open(f"{os.getcwd()}/media/{google_doc_id}.csv", "wb") as f:
        f.write(response.content)

    upload_csv_file(
        write_file_path=f"{os.getcwd()}/media/{google_doc_id}.csv",
        file_name=file_name,
    )

    os.remove(f"{os.getcwd()}/media/{google_doc_id}.csv")


def upload_ogc_api_feature_collection(url: str):
    """
    Downloads a OGC API feature collection and imports it into a PostgreSQL database.

    The collection is downloaded using the requests library and imported into a PostgreSQL database using the psycopg2 library.

    Args:
        url (str): The URL of the OGC API feature collection to be imported.
    """
    collection_id = url.split("collections/")[1].split("/")[0]
    response = requests.get(f"{url}/items")

    if response.status_code != 200:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="There was an error downloading the feature collection",
        )

    feature_collection = response.json()

    numberMatched = feature_collection["numberMatched"]
    numberGenerated = feature_collection["numberReturned"]

    while numberMatched > numberGenerated:
        response = requests.get(f"{url}/items?offset={numberGenerated}")
        feature_collection["features"].extend(response.json()["features"])
        numberGenerated += response.json()["numberReturned"]

    with open(f"{os.getcwd()}/media/{collection_id}.geojson", "w") as f:
        json.dump(feature_collection, f)

    upload_geographic_file(
        file_path=f"{os.getcwd()}/media/{collection_id}.geojson",
        table_name=collection_id,
    )

    os.remove(f"{os.getcwd()}/media/{collection_id}.geojson")


def clean_string(string: str):
    """
    Clean a string by removing spaces, dashes, periods, and colons, and
    converting to lowercase. This is used to clean strings that are used
    as database table names.

    Args:
        string (str): The string to be cleaned.

    Returns:
        str: The cleaned string.
    """
    return (
        string.replace(" ", "_")
        .replace("-", "_")
        .replace(".", "")
        .replace(":", "_")
        .lower()
    )


def upload_ogc_wfs(url: str):
    """
    Downloads an OGC WFS feature collection and imports it into a PostgreSQL database.

    This function retrieves features from an OGC WFS service in batches of 50,
    converting them into a GeoJSON file, which is then imported into a PostgreSQL
    database table. The table name is derived from the typeName parameter in the URL.

    Args:
        url (str): The URL of the OGC WFS service from which to download features.

    Raises:
        HTTPException: If there is an error downloading the WFS feature collection.
    """
    response = requests.get(f"{url}&maxFeatures=50&outputFormat=application%2Fjson")
    collection_id = clean_string(url.split("typeName=")[1].split("&")[0])

    if response.status_code != 200:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="There was an error downloading the wfs",
        )

    feature_collection = response.json()
    total_features = 50
    more_features = True

    while more_features:
        response = requests.get(
            f"{url}&maxFeatures=50&startIndex={total_features}&outputFormat=application%2Fjson"
        )
        if response.json()["features"] == []:
            more_features = False
        feature_collection["features"].extend(response.json()["features"])
        total_features += 50

    with open(f"{os.getcwd()}/media/{collection_id}.geojson", "w") as f:
        json.dump(feature_collection, f)

    upload_geographic_file(
        file_path=f"{os.getcwd()}/media/{collection_id}.geojson",
        table_name=collection_id,
    )

    os.remove(f"{os.getcwd()}/media/{collection_id}.geojson")


def download_data_from_url(url: str):
    """
    Downloads data from a URL and imports it into a PostgreSQL database.

    This function is a wrapper around the requests library and the
    upload_geographic_file function. It downloads a file from the given
    URL and then uploads it to the PostgreSQL database as a new table.

    Args:
        url (str): The URL of the file to be downloaded and imported.

    Raises:
        HTTPException: If there is an error downloading the file.
    """
    response = requests.get(url)

    if response.status_code != 200:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="There was an error downloading the file",
        )

    with open(f"{os.getcwd()}/media/{url.split('/')[-1]}", "wb") as f:
        f.write(response.content)

    upload_geographic_file(
        file_path=f"{os.getcwd()}/media/{url.split('/')[-1]}",
        table_name=url.split("/")[-1].split(".")[0],
    )

    os.remove(f"{os.getcwd()}/media/{url.split('/')[-1]}")


@app.post(path="/api/v1/upload_url")
async def upload_url(url: str):
    """
    Upload a data to the server and import it into a PostgreSQL database.

    Google Sheets
    ESRI Feature Service
    ESRI Map Service
    OGC WFS
    OGC API - Features
    Flat Files
    """

    if "arcgis" in url:
        download_arcgis_service_information(url=url)

    elif "docs.google.com/spreadsheets" in url:
        upload_google_sheets(url=url)

    elif "collection" in url:
        upload_ogc_api_feature_collection(url=url)

    elif "service=wfs" in url.lower():
        upload_ogc_wfs(url=url)

    else:
        download_data_from_url(url=url)

    return {"status": "success"}
